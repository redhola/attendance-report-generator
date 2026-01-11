import pandas as pd
from openpyxl import load_workbook
import re
import datetime
import logging
import os

# Logging configuration for professional monitoring
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

class AttendanceProcessor:
    """
    Processes staff attendance data and generates individual reports based on a template.
    """
    
    # Configuration Constants (Prevents hardcoding inside functions)
    TEMPLATE_NAME = "taslak.xlsx"
    SOURCE_DATA = "DATA.xlsx"
    START_ROW = 6
    END_ROW = 45
    
    def __init__(self, source_path: str, template_path: str):
        self.source_path = source_path
        self.template_path = template_path

    @staticmethod
    def clean_staff_name(name: str) -> str:
        """Removes prefixes like 'arge*' or '*' from staff names."""
        if not isinstance(name, str):
            return str(name)
        # Removes everything up to and including '*' or '-'
        cleaned = re.sub(r'^.*(\*|-)', '', name).strip()
        return cleaned

    def read_source_data(self) -> pd.DataFrame:
        """Reads and pre-processes attendance data from Excel."""
        try:
            wb = load_workbook(self.source_path, data_only=True)
            ws = wb.active
            data_rows = []

            for row in ws.iter_rows(min_row=5, values_only=True):
                # Using indexed access for clarity based on your Excel structure
                # B=1, G=6, H=7, J=9, M=12
                if not row[6]:  # Skip if date is empty
                    continue

                data_rows.append({
                    "Staff": self.clean_staff_name(row[1]),
                    "DateRaw": row[6],
                    "Entry": row[7],
                    "Exit": row[9],
                    "NetDuration": row[12]
                })

            df = pd.DataFrame(data_rows)
            df["Date"] = pd.to_datetime(df["DateRaw"], dayfirst=True, errors="coerce").dt.date
            return df.dropna(subset=["Date"])
        
        except Exception as e:
            logging.error(f"Error reading source file: {e}")
            return pd.DataFrame()

    def create_daily_summary(self, staff_df: pd.DataFrame) -> pd.DataFrame:
        """Aggregates multiple entries/exits into a single daily summary."""
        daily_data = []
        grouped = staff_df.groupby("Date")

        for date_val, group in grouped:
            group_sorted = group.sort_values(by="DateRaw")
            
            entry_time = group_sorted.iloc[0]["Entry"]
            # Logic: Last exit is often the penultimate row or the last one
            exit_time = group_sorted.iloc[-2]["Exit"] if len(group_sorted) > 1 else group_sorted.iloc[-1]["Exit"]
            net_duration = group_sorted.iloc[-1]["NetDuration"]

            daily_data.append({
                "Date": date_val,
                "EntryTime": entry_time,
                "ExitTime": exit_time,
                "NetDuration": net_duration
            })

        return pd.DataFrame(daily_data)

    def fill_template(self, staff_name: str, daily_df: pd.DataFrame):
        """Populates the Excel template for a specific staff member."""
        try:
            wb = load_workbook(self.template_path)
            ws = wb.active
            ws["F4"] = staff_name  # Set Staff Name

            # Helper to write time objects to cells
            def _write_time_cell(cell, time_val):
                if time_val and isinstance(time_val, str):
                    try:
                        cell.value = datetime.datetime.strptime(time_val, "%H:%M:%S").time()
                        cell.number_format = "hh:mm:ss"
                    except ValueError:
                        cell.value = time_val
                else:
                    cell.value = time_val

            for row_idx in range(self.START_ROW, self.END_ROW):
                cell_date_val = ws[f"E{row_idx}"].value
                if not cell_date_val:
                    continue
                
                # Standardize date format for comparison
                current_row_date = pd.to_datetime(cell_date_val).date()
                match = daily_df[daily_df["Date"] == current_row_date]

                if not match.empty:
                    row_data = match.iloc[0]
                    _write_time_cell(ws[f"F{row_idx}"], row_data["EntryTime"])
                    _write_time_cell(ws[f"G{row_idx}"], row_data["ExitTime"])
                    _write_time_cell(ws[f"I{row_idx}"], row_data["NetDuration"])

            output_file = f"{staff_name.replace(' ', '_')}_Attendance.xlsx"
            wb.save(output_file)
            logging.info(f"Report generated: {output_file}")

        except Exception as e:
            logging.error(f"Error filling template for {staff_name}: {e}")

    def run(self):
        """Main execution flow."""
        df_all = self.read_source_data()
        if df_all.empty:
            logging.warning("No data found to process.")
            return

        # Filter out invalid staff names
        excluded_keywords = ["toplam", "günlük", "personel"]
        unique_staff = [
            s for s in df_all["Staff"].unique() 
            if s and not any(k in s.lower() for k in excluded_keywords)
        ]

        for person in unique_staff:
            person_df = df_all[df_all["Staff"] == person]
            daily_summary = self.create_daily_summary(person_df)
            self.fill_template(person, daily_summary)

if __name__ == "__main__":
    processor = AttendanceProcessor("DATA.xlsx", "taslak.xlsx")
    processor.run()